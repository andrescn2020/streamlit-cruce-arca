import streamlit as st
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import NamedStyle
import zipfile


# ============================================================================
# FUNCIONES DE LECTURA Y LIMPIEZA DE ARCHIVOS
# ============================================================================


def leer_archivo(file_path):
    """Lee el archivo con manejo de codificación UTF-8 o Latin-1"""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.readlines()
    except UnicodeDecodeError:
        with open(file_path, "r", encoding="latin-1") as f:
            return f.readlines()


def procesar_encabezado(lines):
    """Procesa y extrae la información del encabezado del archivo"""
    try:
        encabezado = lines[1:7]
        encabezado_limpio = [line.replace("\n", "").strip() for line in encabezado]

        return {
            "RAZON SOCIAL": encabezado_limpio[0],
            "DIRECCION": encabezado_limpio[1],
            "CUIT": encabezado_limpio[2],
            "LIBRO": encabezado_limpio[3].split("  ")[-1],
            "PERIODO": encabezado_limpio[4].split("  ")[-1],
        }
    except Exception as e:
        st.error(f"Ocurrió un error al procesar el encabezado: {e}")
        return {}


def limpiar_lineas(lines):
    """Limpia las líneas del archivo eliminando caracteres de control y bloques no deseados"""
    cleaned_lines = []
    eliminar = False
    eliminar_desde_totales = False
    compras_o_ventas = ""

    for i, line in enumerate(lines[9:], start=2):
        # Detectar tipo de operación
        if "IVA VENTAS" in line:
            compras_o_ventas = "Ventas"
        elif "IVA COMPRAS" in line:
            compras_o_ventas = "Compras"

        # Detectar fin de datos
        if "TOTALES POR TASA" in line:
            eliminar_desde_totales = True
            continue

        if eliminar_desde_totales:
            break

        # Manejar bloques a eliminar
        if line.startswith("----"):
            eliminar = True
            continue

        if line.startswith("--"):
            eliminar = False
            continue

        # Procesar líneas válidas
        if not eliminar:
            cleaned_line = re.sub(r"\x1b[^m]*m", "", line)  # Elimina secuencias ANSI
            cleaned_line = re.sub(
                r"[\x00-\x1F\x7F]", "", cleaned_line
            )  # Eliminación de caracteres de control ASCII
            cleaned_lines.append(cleaned_line)

    return cleaned_lines, compras_o_ventas


def limpiar_lineas_adicional(cleaned_lines):
    """Segunda limpieza de líneas eliminando líneas con PPag"""
    doble_cleaned_lines = []

    for i, line in enumerate(cleaned_lines, start=-1):
        if "PPag." in line or len(line.strip()) < 35:
            if re.search(r"PPag\.\:\s*\d+\s*$", line):
                linea = re.sub(r"PPag\.\:\s*\d+\s*$", "", line)
                cleaned_lines.append(linea)
            else:
                break
        doble_cleaned_lines.append(line)

    return doble_cleaned_lines


# ============================================================================
# FUNCIONES DE PROCESAMIENTO DE MOVIMIENTOS
# ============================================================================


def procesar_movimientos(doble_cleaned_lines, compras_o_ventas):
    """Procesa las líneas limpias y extrae los movimientos"""
    movements = []
    temp_movement = {}

    for index, cleaned_line in enumerate(doble_cleaned_lines):
        # Procesar líneas continuas del mismo movimiento
        if "numero" in temp_movement and temp_movement["numero"] == cleaned_line[12:20]:
            procesar_linea_continuacion(cleaned_line, temp_movement, compras_o_ventas)
        else:
            # Procesar nueva línea de movimiento
            if cleaned_line[0:2] == "  ":
                procesar_linea_continuacion(
                    cleaned_line, temp_movement, compras_o_ventas
                )
                if index == len(doble_cleaned_lines) - 1:
                    movements.append(temp_movement)
            else:
                # Nueva entrada de movimiento
                movement = temp_movement.copy()
                movements.append(movement)
                temp_movement.clear()

                procesar_nueva_entrada(cleaned_line, temp_movement, compras_o_ventas)

                if index == len(doble_cleaned_lines) - 1:
                    movements.append(temp_movement)

    # Limpiar movimiento vacío inicial
    if not movements[0]:
        movements.pop(0)

    return movements


def procesar_linea_continuacion(cleaned_line, temp_movement, compras_o_ventas):
    """Procesa una línea que continúa un movimiento existente"""
    partes = re.split(r"\s{3,}", cleaned_line[70:])
    if len(partes) < 2:
        return

    tasa = partes[0]
    if tasa in [
        "Tasa 21%",
        "T.10.5%",
        "Tasa 27%",
        "C.F.21%",
        "C.F.10.5%",
        "Tasa 2.5%",
        "T.IMP 21%",
        "T.IMP 10%",
    ]:
        procesar_tasa_con_neto_iva(tasa, partes, temp_movement)
    elif tasa in ["R.Monot21", "R.Mont.10"]:
        procesar_tasa_monotributo(tasa, partes, temp_movement, compras_o_ventas)
    else:
        procesar_otra_tasa(tasa, partes, temp_movement)


def procesar_tasa_con_neto_iva(tasa, partes, temp_movement):
    """Procesa tasas que tienen neto e IVA separados"""
    if tasa + " Neto" in temp_movement:
        # Sumar valores existentes
        neto_anterior = float(temp_movement[tasa + " Neto"].replace(",", "."))
        iva_anterior = float(temp_movement[tasa + " IVA"].replace(",", "."))
        neto_nuevo = float(partes[1].replace(",", "."))
        iva_nuevo = float(partes[2].replace(",", "."))

        temp_movement[tasa + " Neto"] = round(neto_anterior + neto_nuevo, 2)
        temp_movement[tasa + " IVA"] = round(iva_anterior + iva_nuevo, 2)
    else:
        # Primer valor
        temp_movement[tasa + " Neto"] = partes[1]
        temp_movement[tasa + " IVA"] = partes[2]


def procesar_tasa_monotributo(tasa, partes, temp_movement, compras_o_ventas):
    """Procesa tasas de monotributo"""
    if compras_o_ventas == "Ventas":
        if tasa + " Neto" in temp_movement:
            neto_anterior = float(temp_movement[tasa + " Neto"].replace(",", "."))
            iva_anterior = float(temp_movement[tasa + " IVA"].replace(",", "."))
            neto_nuevo = float(partes[1].replace(",", "."))
            iva_nuevo = float(partes[2].replace(",", "."))

            temp_movement[tasa + " Neto"] = round(neto_anterior + neto_nuevo, 2)
            temp_movement[tasa + " IVA"] = round(iva_anterior + iva_nuevo, 2)
        else:
            temp_movement[tasa + " Neto"] = partes[1]
            temp_movement[tasa + " IVA"] = partes[2]
    else:
        temp_movement[tasa] = partes[1]


def procesar_otra_tasa(tasa, partes, temp_movement):
    """Procesa otras tasas que no tienen neto/IVA separados"""
    if tasa in temp_movement:
        if isinstance(temp_movement[tasa], float):
            numero_actual = float(partes[1].replace(",", "."))
            temp_movement[tasa] = round(temp_movement[tasa] + numero_actual, 2)
        else:
            numero_anterior = float(temp_movement[tasa].replace(",", "."))
            numero_actual = float(partes[1].replace(",", "."))
            temp_movement[tasa] = round(numero_anterior + numero_actual, 2)
    else:
        temp_movement[tasa] = partes[1]


def procesar_nueva_entrada(cleaned_line, temp_movement, compras_o_ventas):
    """Procesa una nueva entrada de movimiento"""
    partes = re.split(r"\s{3,}", cleaned_line[70:])
    if len(partes) < 2:
        return

    temp_movement.update(
        {
            "Fecha": cleaned_line[0:2],
            "Comprobante": cleaned_line[3:5],
            "PV": cleaned_line[6:11],
            "Nro": cleaned_line[12:20],
            "Letra": cleaned_line[20:21],
            "Razon Social": cleaned_line[22:44],
            "Condicion": cleaned_line[45:49],
            "CUIT": cleaned_line[50:63],
            "Concepto": cleaned_line[64:67],
            "Jurisdiccion": cleaned_line[68:69],
        }
    )

    # Procesar montos
    if len(partes) == 3:
        primer_monto = partes[1].split(" ")
        primer_monto = list(filter(None, primer_monto))
        segundo_monto = partes[1].split(" ")
        segundo_monto = list(filter(None, segundo_monto))
        partes = [partes[0]] + primer_monto + segundo_monto

    tasa = partes[0]
    if tasa in [
        "Tasa 21%",
        "T.10.5%",
        "Tasa 27%",
        "C.F.21%",
        "C.F.10.5%",
        "Tasa 2.5%",
        "T.IMP 21%",
        "T.IMP 10%",
    ]:
        temp_movement[tasa + " Neto"] = partes[1]
        temp_movement[tasa + " IVA"] = partes[2]
    elif tasa in ["R.Monot21", "R.Mont.10"]:
        if compras_o_ventas == "Ventas":
            temp_movement[tasa + " Neto"] = partes[1]
            temp_movement[tasa + " IVA"] = partes[2]
        else:
            temp_movement[tasa] = partes[1]
    else:
        temp_movement[tasa] = partes[1]


# ============================================================================
# FUNCIONES DE PROCESAMIENTO DE DATAFRAMES
# ============================================================================


def crear_dataframe_movimientos(movements):
    """Crea y procesa el DataFrame de movimientos"""
    df = pd.DataFrame(movements)
    df = df.fillna(0)

    # Reemplazar comas por puntos en columnas numéricas
    df.iloc[:, 11:] = df.iloc[:, 11:].replace(",", ".", regex=True)
    df.iloc[:, 11:] = df.iloc[:, 11:].apply(pd.to_numeric, errors="coerce").fillna(0)

    # Convertir notas de crédito a negativas
    columnas_a_convertir = df.columns[11:]
    df.loc[df["Comprobante"] == "NC", columnas_a_convertir] *= -1

    # Convertir tipos de datos
    df["PV"] = pd.to_numeric(df["PV"])
    df["Nro"] = pd.to_numeric(df["Nro"])
    df["Concepto"] = pd.to_numeric(df["Concepto"])

    return df


def combinar_movimientos_duplicados(df):
    """Combina movimientos que tienen la misma clave principal"""
    resultado = []
    fila_actual = df.iloc[0].copy()

    for i in range(1, len(df)):
        fila_siguiente = df.iloc[i]

        # Si la clave principal se repite, combinar valores
        if (
            fila_actual["Nro"] == fila_siguiente["Nro"]
            and fila_actual["PV"] == fila_siguiente["PV"]
            and fila_actual["Razon Social"] == fila_siguiente["Razon Social"]
        ):

            for col in df.columns[11:]:  # Sumar solo las columnas numéricas
                fila_actual[col] += fila_siguiente[col]
        else:
            resultado.append(fila_actual)
            fila_actual = fila_siguiente.copy()

    resultado.append(fila_actual)
    return pd.DataFrame(resultado)


def agregar_totales_movimientos(df_final):
    """Agrega fila de totales al DataFrame de movimientos"""
    df_final["Total"] = df_final.iloc[:, 11:].sum(axis=1)

    # Crear fila de totales
    fila_total = pd.DataFrame(df_final.iloc[:, 11:].sum()).T
    fila_total.insert(0, "Nro", "TOTALES")
    fila_total.insert(1, "Razon Social", "")

    return pd.concat([df_final, fila_total], ignore_index=True)


# ============================================================================
# FUNCIONES DE EXCEL
# ============================================================================


def crear_archivo_excel_consolidado(
    df_mendez, df_arca, df_arca_no_en_mendez, df_mendez_no_en_arca
):
    """Crea el archivo Excel consolidado con 4 hojas"""
    excel_filename = "Cruce_Consolidado.xlsx"

    with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
        # Hoja 1: Mendez (movimientos del TXT)
        df_mendez.to_excel(writer, sheet_name="Mendez", index=False)

        # Hoja 2: ARCA (movimientos del ZIP)
        df_arca.to_excel(writer, sheet_name="ARCA", index=False)

        # Hoja 3: ARCA NO EN MENDEZ (comprobantes en ARCA y no en Mendez)
        df_arca_no_en_mendez.to_excel(
            writer, sheet_name="ARCA NO EN MENDEZ", index=False
        )

        # Hoja 4: MENDEZ NO EN ARCA (comprobantes en Mendez y no en ARCA)
        df_mendez_no_en_arca.to_excel(
            writer, sheet_name="MENDEZ NO EN ARCA", index=False
        )

    return excel_filename


def aplicar_formato_excel_consolidado(excel_filename):
    """Aplica formato de moneda a todas las hojas del archivo consolidado"""
    wb = load_workbook(excel_filename)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Encontrar la última fila y columna con datos
        max_row = ws.max_row
        max_col = ws.max_column

        # Aplicar formato de moneda a todas las columnas numéricas (desde la columna 11 en adelante)
        for col_idx in range(11, max_col + 1):
            col_letter = get_column_letter(col_idx)
            for row_idx in range(
                2, max_row + 1
            ):  # Empezar desde la fila 2 (después del encabezado)
                cell = ws[f"{col_letter}{row_idx}"]
                if cell.value is not None:
                    cell.number_format = '"$"#,##0.00'

    wb.save(excel_filename)


def crear_archivo_excel(df_encabezado, df_final):
    """Crea el archivo Excel solo con la hoja de movimientos"""
    excel_filename = "Movimientos.xlsx"

    with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
        # Solo hoja Movimientos
        df_encabezado.to_excel(
            writer, sheet_name="Movimientos", startcol=5, index=False
        )
        df_final.to_excel(writer, sheet_name="Movimientos", startrow=8, index=False)

    return excel_filename


def aplicar_formulas_excel(excel_filename, df_final):
    """Aplica formato de moneda en la hoja de movimientos, sin agregar ninguna fórmula de suma."""
    wb = load_workbook(excel_filename)
    wm = wb["Movimientos"]

    # Calcular la fila donde empieza la tabla en Excel
    inicio_fila_movimientos = 9
    ultima_fila_movimientos = inicio_fila_movimientos + df_final.shape[0] - 1

    # Aplicar formato de moneda a todas las columnas numéricas (desde columna 11 en adelante)
    for col_idx in range(11, df_final.shape[1] + 1):
        col_letter = get_column_letter(col_idx)
        for row_idx in range(
            inicio_fila_movimientos + 1, ultima_fila_movimientos + 2
        ):  # +2 para incluir la última fila
            cell = wm[f"{col_letter}{row_idx}"]
            cell.number_format = '"$"#,##0.00'

    wb.save(excel_filename)


# ============================================================================
# FUNCIÓN PRINCIPAL
# ============================================================================


def procesar_archivo(file_path):
    """Función principal que procesa el archivo completo"""
    try:
        # 1. Leer y limpiar archivo
        lines = leer_archivo(file_path)
        encabezado_completo = procesar_encabezado(lines)
        cleaned_lines, compras_o_ventas = limpiar_lineas(lines)
        doble_cleaned_lines = limpiar_lineas_adicional(cleaned_lines)

        # 2. Procesar movimientos
        movements = procesar_movimientos(doble_cleaned_lines, compras_o_ventas)

        # 3. Crear DataFrames
        df = crear_dataframe_movimientos(movements)
        df_final = combinar_movimientos_duplicados(df)
        df_final = agregar_totales_movimientos(df_final)

        # Forzar tipo numérico en columnas desde la 11 en adelante
        for col in df_final.columns[11:]:
            df_final[col] = (
                pd.to_numeric(df_final[col], errors="coerce").fillna(0).astype(float)
            )

        # 4. Preparar DataFrame de encabezado
        df_encabezado = pd.DataFrame(
            list(encabezado_completo.values()), columns=["Valor"]
        )
        df_encabezado.columns = [""] * len(df_encabezado.columns)

        # 5. Crear archivo Excel (sin la fila de totales)
        df_final_sin_totales = df_final[df_final["Nro"] != "TOTALES"].copy()

        # Forzar tipo float en todas las columnas numéricas posteriores a 'Jurisdiccion'
        idx_jurisdiccion = df_final_sin_totales.columns.get_loc("Jurisdiccion")
        for col in df_final_sin_totales.columns[idx_jurisdiccion + 1 :]:
            df_final_sin_totales[col] = (
                df_final_sin_totales[col]
                .astype(str)
                .str.replace(",", ".", regex=False)
                .str.replace(" ", "", regex=False)
            )
            df_final_sin_totales[col] = (
                pd.to_numeric(df_final_sin_totales[col], errors="coerce")
                .fillna(0)
                .astype(float)
            )

        excel_filename = crear_archivo_excel(df_encabezado, df_final_sin_totales)

        # 6. Aplicar formato
        aplicar_formulas_excel(excel_filename, df_final_sin_totales)

        st.success("¡Archivo procesado con éxito!")
        return excel_filename, df_final_sin_totales

    except Exception as e:
        st.error(f"Error al procesar el archivo: {e}")
        return None, None


# ============================================================================
# INTERFAZ DE STREAMLIT
# ============================================================================


def procesar_zip_csv(zip_path):
    with zipfile.ZipFile(zip_path, "r") as z:
        for file in z.namelist():
            if file.endswith(".csv"):
                with z.open(file) as f:
                    df_zip = pd.read_csv(f, dtype=str, delimiter=";")
                return df_zip
    return None


def comprobantes_faltantes(df_txt, df_zip):
    def limpiar_numero(val, largo):
        if pd.isnull(val):
            return "".zfill(largo)
        try:
            val = int(float(val))
        except Exception:
            pass
        return str(val).zfill(largo)

    df_txt["PV"] = df_txt["PV"].apply(lambda x: limpiar_numero(x, 5))
    df_txt["Nro"] = df_txt["Nro"].apply(lambda x: limpiar_numero(x, 8))
    df_zip["Punto de Venta"] = df_zip["Punto de Venta"].apply(
        lambda x: limpiar_numero(x, 5)
    )
    df_zip["Número de Comprobante"] = df_zip["Número de Comprobante"].apply(
        lambda x: limpiar_numero(x, 8)
    )
    df_txt["clave"] = df_txt["PV"] + "-" + df_txt["Nro"]
    df_zip["clave"] = df_zip["Punto de Venta"] + "-" + df_zip["Número de Comprobante"]
    faltantes = df_zip[~df_zip["clave"].isin(df_txt["clave"])].copy()
    if "clave" in faltantes.columns:
        faltantes = faltantes.drop(columns=["clave"])
    return faltantes


def comprobantes_faltantes_inverso(df_txt, df_zip):
    def limpiar_numero(val, largo):
        if pd.isnull(val):
            return "".zfill(largo)
        try:
            val = int(float(val))
        except Exception:
            pass
        return str(val).zfill(largo)

    df_txt["PV"] = df_txt["PV"].apply(lambda x: limpiar_numero(x, 5))
    df_txt["Nro"] = df_txt["Nro"].apply(lambda x: limpiar_numero(x, 8))
    df_zip["Punto de Venta"] = df_zip["Punto de Venta"].apply(
        lambda x: limpiar_numero(x, 5)
    )
    df_zip["Número de Comprobante"] = df_zip["Número de Comprobante"].apply(
        lambda x: limpiar_numero(x, 8)
    )
    df_txt["clave"] = df_txt["PV"] + "-" + df_txt["Nro"]
    df_zip["clave"] = df_zip["Punto de Venta"] + "-" + df_zip["Número de Comprobante"]
    faltantes = df_txt[~df_txt["clave"].isin(df_zip["clave"])].copy()
    if "clave" in faltantes.columns:
        faltantes = faltantes.drop(columns=["clave"])
    return faltantes


def main():
    st.set_page_config(
        page_title="Procesador de Movimientos IVA", page_icon="📊", layout="wide"
    )

    st.title("📊 Procesador de Movimientos IVA")
    st.markdown("---")

    # Subir archivo TXT
    uploaded_file = st.file_uploader(
        "Selecciona el archivo de movimientos IVA",
        type=["txt"],
        help="Sube un archivo de texto con los movimientos IVA",
    )

    # Subir archivo ZIP
    uploaded_zip = st.file_uploader(
        "Selecciona el archivo ZIP adicional",
        type=["zip"],
        help="Sube un archivo ZIP adicional",
    )

    if uploaded_file is None or uploaded_zip is None:
        st.warning("Debes subir ambos archivos (TXT y ZIP) para continuar.")
        st.stop()

    # Guardar archivos temporalmente
    with open("temp_file.txt", "wb") as f:
        f.write(uploaded_file.getbuffer())
    with open("temp_file.zip", "wb") as f:
        f.write(uploaded_zip.getbuffer())

    with st.spinner("Procesando archivos..."):
        # Procesar archivo TXT
        excel_filename, df_mendez = procesar_archivo("temp_file.txt")

        # Procesar archivo ZIP
        df_arca = procesar_zip_csv("temp_file.zip")

        if df_arca is not None:
            # Crear DataFrames para el cruce
            faltantes_arca_no_mendez = comprobantes_faltantes(df_mendez, df_arca)
            faltantes_mendez_no_arca = comprobantes_faltantes_inverso(
                df_mendez, df_arca
            )

            # Crear archivo Excel consolidado
            excel_consolidado = crear_archivo_excel_consolidado(
                df_mendez, df_arca, faltantes_arca_no_mendez, faltantes_mendez_no_arca
            )

            # Aplicar formato al archivo consolidado
            aplicar_formato_excel_consolidado(excel_consolidado)

            st.success("✅ Archivos procesados correctamente!")

            # Solo un botón de descarga para el archivo consolidado
            with open(excel_consolidado, "rb") as f:
                st.download_button(
                    label="📥 Descargar Excel Consolidado",
                    data=f.read(),
                    file_name=excel_consolidado,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        else:
            st.error("No se encontró un archivo CSV dentro del ZIP.")
            # Limpiar archivos temporales
            import os

            if os.path.exists("temp_file.txt"):
                os.remove("temp_file.txt")
            if os.path.exists("temp_file.zip"):
                os.remove("temp_file.zip")
            st.stop()

    # Limpiar archivos temporales
    import os

    if os.path.exists("temp_file.txt"):
        os.remove("temp_file.txt")
    if os.path.exists("temp_file.zip"):
        os.remove("temp_file.zip")
    if os.path.exists("Movimientos.xlsx"):
        os.remove("Movimientos.xlsx")
    if os.path.exists("Comprobantes_faltantes_zip_no_txt.xlsx"):
        os.remove("Comprobantes_faltantes_zip_no_txt.xlsx")
    if os.path.exists("Comprobantes_faltantes_txt_no_zip.xlsx"):
        os.remove("Comprobantes_faltantes_txt_no_zip.xlsx")


if __name__ == "__main__":
    main()
